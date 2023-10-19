# Programa que faz automacao de portateis Mac do site da PCBEM
# Este programa colocara todos os portateis numa folha Excel juntamente com o seu preço

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl

# Acessar o site
driver = webdriver.Edge()
driver.get("https://www.pcbem.pt/loja/categoria/computadores/portateis/apple-portateis/")

# WebDriverWait para esperar que os elementos sejam visíveis
wait = WebDriverWait(driver, 10)  # Tempo limite (10 segundos)

# Esperar até que pelo menos um elemento de título seja visível
wait.until(EC.presence_of_element_located((By.XPATH, "//h2[@class='woocommerce-loop-product__title']")))

# Esperar até que pelo menos um elemento de preço seja visível
wait.until(EC.presence_of_element_located((By.XPATH, "//span[@class='woocommerce-Price-amount amount']")))

# extrair todos os títulos
titulos = driver.find_elements(By.XPATH, "//h2[@class='woocommerce-loop-product__title']")

# extrair todos os preços
precos = driver.find_elements(By.XPATH, "//span[@class='woocommerce-Price-amount amount']")

# Criando um livro Excel
workBook = openpyxl.Workbook()

# Criando um Folha com nome 'produtos'
workBook.create_sheet('produtos')

# Seleciono a pagina guardando-a numa variável
folha_produto = workBook['produtos']

# Registar 'Produto' na célula A1 e 'Preço' na célula B2
folha_produto['A1'].value = "Produto"
folha_produto['B1'].value = "Preço"

# inserir os titulos e precos no execel
for x in range(len(titulos)):
    folha_produto.cell(row=x + 2, column=1, value=titulos[x].text)
    folha_produto.cell(row=x + 2, column=2, value=precos[x].text)

# Guardando o livro Excel
workBook.save('produtos.xlsx')

# Fechar o navegador
driver.quit()